import openpyxl
import re
from collections import defaultdict

# Abrir el archivo
wb = openpyxl.load_workbook(r"C:\Users\IT\Desktop\Productos Chapalita.xlsx", read_only=True, data_only=True)
ws = wb.active

# Buscar la columna de ItemCode y ItemName
header_row = None
code_col = None
name_col = None

for row in ws.iter_rows(min_row=1, max_row=10):
    for cell in row:
        val = str(cell.value).strip().lower() if cell.value else ""
        if val in ("itemcode", "item code", "codigo", "código", "code"):
            code_col = cell.column
            header_row = cell.row
        elif val in ("itemname", "item name", "nombre", "name", "descripcion", "descripción"):
            name_col = cell.column
            header_row = cell.row

if not code_col:
    # Intentar primera columna
    print("No encontré header, usando columna A y B")
    code_col = 1
    name_col = 2
    header_row = 0

print(f"Header en fila {header_row}, Code col={code_col}, Name col={name_col}")

# Extraer prefijo: parte alfabética al inicio del código
def extraer_prefijo(codigo):
    if not codigo:
        return None
    codigo = str(codigo).strip()
    match = re.match(r'^([A-Za-z]+)', codigo)
    if match:
        return match.group(1).upper()
    return None

# Procesar filas
prefijos = defaultdict(lambda: {"count": 0, "ejemplos": []})
total = 0
sin_prefijo = 0

for row in ws.iter_rows(min_row=(header_row or 0) + 1, values_only=False):
    code_val = row[code_col - 1].value if code_col and len(row) >= code_col else None
    name_val = row[name_col - 1].value if name_col and len(row) >= name_col else None
    
    if not code_val:
        continue
    
    total += 1
    prefijo = extraer_prefijo(code_val)
    
    if prefijo:
        prefijos[prefijo]["count"] += 1
        if len(prefijos[prefijo]["ejemplos"]) < 2:
            prefijos[prefijo]["ejemplos"].append(f"{str(code_val).strip()} - {str(name_val).strip()[:60] if name_val else '?'}")
    else:
        sin_prefijo += 1

wb.close()

# Ordenar por prefijo
sorted_prefijos = sorted(prefijos.items(), key=lambda x: x[0])

print(f"\nTotal filas procesadas: {total}")
print(f"Prefijos únicos encontrados: {len(prefijos)}")
print(f"Sin prefijo: {sin_prefijo}")
print("\n" + "="*100)
print(f"{'PREFIJO':<12} {'CANT':>5}  {'EJEMPLO'}")
print("="*100)

for pref, data in sorted_prefijos:
    ejemplo = data["ejemplos"][0] if data["ejemplos"] else "?"
    print(f"{pref:<12} {data['count']:>5}  {ejemplo}")

print("="*100)
