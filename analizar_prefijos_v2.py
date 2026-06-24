import openpyxl
import re
from collections import defaultdict

wb = openpyxl.load_workbook(r"C:\Users\IT\Desktop\Productos Chapalita.xlsx", read_only=True, data_only=True)
ws = wb.active

# Detectar headers
code_col = 1
name_col = 2
header_row = 1

for row in ws.iter_rows(min_row=1, max_row=5):
    for cell in row:
        val = str(cell.value).strip().lower() if cell.value else ""
        if val in ("itemcode", "item code", "codigo", "código", "code"):
            code_col = cell.column
            header_row = cell.row
        elif val in ("itemname", "item name", "nombre", "name", "descripcion", "descripción"):
            name_col = cell.column

def extraer_prefijo(codigo):
    if not codigo:
        return None
    codigo = str(codigo).strip()
    match = re.match(r'^([A-Za-z]+)', codigo)
    if match:
        return match.group(1).upper()
    return None

prefijos = defaultdict(lambda: {"count": 0, "ejemplos": []})

for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
    code_val = row[code_col - 1].value if len(row) >= code_col else None
    name_val = row[name_col - 1].value if len(row) >= name_col else None
    if not code_val:
        continue
    prefijo = extraer_prefijo(code_val)
    if prefijo:
        prefijos[prefijo]["count"] += 1
        if len(prefijos[prefijo]["ejemplos"]) < 3:
            prefijos[prefijo]["ejemplos"].append({
                "code": str(code_val).strip(),
                "name": str(name_val).strip()[:80] if name_val else "?"
            })

wb.close()

# Clasificar prefijos
# Productos terminados/comerciales (con estructura E/M/N + familia)
comerciales = {}
materiales = {}
muestras = {}
otros = {}

for pref, data in sorted(prefijos.items()):
    if pref.startswith("MUE"):
        muestras[pref] = data
    elif pref.startswith(("E","M","N")) and len(pref) >= 4 and pref[1:3] in ("AE","AR","BR","DI","LI","NO","PA","PC","TA","RE","CA"):
        comerciales[pref] = data
    elif pref.startswith(("E","M","N")) and len(pref) >= 3 and pref not in ("ME","MI","MM","MO","MP","MPI","MS","MZ"):
        comerciales[pref] = data
    elif pref in ("ME","MP","MPI","MS","RP","FMPI","FMIP","FMIPI","MPIDA","MPIVA","EMPI"):
        materiales[pref] = data
    elif pref in ("BL","CM","CN","HER","HEE","HET","MI","MM","MO","MZ","PCN","PIC","RI","RO","REF","REP","RET","SEGG","TPM"):
        otros[pref] = data
    else:
        otros[pref] = data

# Imprimir resultado
print("=" * 110)
print("PRODUCTOS COMERCIALES (Aerosoles, Aromatizantes, Limpieza, etc.)")
print("=" * 110)
print(f"{'PREFIJO':<14} {'CANT':>5}  {'SIGNIFICADO PROPUESTO':<50} {'EJEMPLO'}")
print("-" * 110)

# Diccionario de significados conocidos
significados = {
    "EAEHO": "Exportación | Aerosol | Home (ambiente/lustrador)",
    "MAEHO": "Maquila | Aerosol | Home (ambiente/lustrador)",
    "NAEHO": "Nacional | Aerosol | Home (ambiente/lustrador)",
    "EAEDC": "Exportación | Aerosol | Dosificado Continuo (Metered)",
    "MAEDC": "Maquila | Aerosol | Dosificado Continuo (Metered)",
    "NAEDC": "Nacional | Aerosol | Dosificado Continuo (Metered)",
    "EAEAC": "Exportación | Aerosol | Air Care (NILODOR/ZENEX)",
    "EAEMC": "Exportación | Aerosol | Micro-Can",
    "NAEMC": "Nacional | Aerosol | Micro-Can",
    "EAEMS": "Exportación | Aerosol | Mini Spray",
    "NAEMS": "Nacional | Aerosol | Mini Spray",
    "EAENI": "Exportación | Aerosol | Non-aerosol/Spray (Room Spray)",
    "EAEHH": "Exportación | Aerosol | Home/Hogar (espuma/carpet)",
    "EAEIN": "Exportación | Aerosol | Insecticida",
    "NAEIN": "Nacional | Aerosol | Insecticida",
    "EACPL": "Exportación | Aerosol | Charola/Plato (Drip Tray)",
    "EACPO": "Exportación | Accesorio | ? (Urinal Screen accesories)",
    "EARCG": "Exportación | Aromatizante | Cono Gel",
    "MARCG": "Maquila | Aromatizante | Cono Gel",
    "NARCG": "Nacional | Aromatizante | Cono Gel",
    "EARGE": "Exportación | Aromatizante | Gel (lata 70g)",
    "MARGE": "Maquila | Aromatizante | Gel (lata 70g)",
    "NARGE": "Nacional | Aromatizante | Gel (lata 70g)",
    "EARCO": "Exportación | Aromatizante | Clip-On (Bowl Clip/EVA)",
    "NARCO": "Nacional | Aromatizante | Clip-On",
    "EARHA": "Exportación | Aromatizante | Hang Air (colgante)",
    "NARHA": "Nacional | Aromatizante | Hang Air",
    "NARAU": "Nacional | Aromatizante | Auto",
    "NREAU": "Nacional | Repuesto | Auto",
    "EBRDR": "Exportación | Breeze Matic | Dispensador/Kit",
    "MBRDR": "Maquila | Breeze Matic | Dispensador",
    "NBRDR": "Nacional | Breeze Matic | Dispensador/Kit",
    "EDIDC": "Exportación | Dispensador | Dosificado Continuo",
    "MDIDC": "Maquila | Dispensador | Dosificado Continuo",
    "NDIDC": "Nacional | Dispensador | Dosificado Continuo",
    "EDIEL": "Exportación | Difusor Eléctrico | Kit",
    "MDIEL": "Maquila | Difusor Eléctrico | Kit",
    "NDIEL": "Nacional | Difusor Eléctrico | Kit",
    "MDILE": "Maquila | Difusor Eléctrico (typo de MDIEL)",
    "EDIER": "Exportación | Difusor Eléctrico | Repuesto",
    "MDIER": "Maquila | Difusor Eléctrico | Repuesto",
    "NDIER": "Nacional | Difusor Eléctrico | Repuesto",
    "EDILG": "Exportación | Dispensador | Líquido Goteador",
    "NDILG": "Nacional | Dispensador | Líquido Goteador",
    "ELILG": "Exportación | Líquido | Líquido Goteador (refill/conc.)",
    "NLILG": "Nacional | Líquido | Líquido Goteador",
    "ENOCA": "Exportación | Sanitario | Canastilla",
    "MNOCA": "Maquila | Sanitario | Canastilla",
    "NNOCA": "Nacional | Sanitario | Canastilla",
    "ENOCG": "Exportación | Sanitario | Canastilla Gel",
    "NNOCG": "Nacional | Sanitario | Canastilla Gel",
    "NNORC": "Nacional | Sanitario | Repuesto Canastilla",
    "ENOCR": "Exportación | Sanitario | Cristales NON-PDCB",
    "NNOCR": "Nacional | Sanitario | Cristales NON-PDCB",
    "ENOPO": "Exportación | Sanitario | Porta Potty",
    "ENOPA": "Exportación | Sanitario | Pastilla/Accesorio (varios)",
    "MNOPA": "Maquila | Sanitario | Pastilla Azul",
    "NNOPA": "Nacional | Sanitario | Pastilla Azul",
    "NNOMI": "Nacional | Sanitario | Mingitorio (pastilla)",
    "EPABA": "Exportación | Pastilla | Barra (Wall Block)",
    "NPABA": "Nacional | Pastilla | Barra",
    "EPAAL": "Exportación | Pastilla | Alambre",
    "MPAAL": "Maquila | Pastilla | Alambre",
    "NPAAL": "Nacional | Pastilla | Alambre",
    "EPAMI": "Exportación | Pastilla | Mingitorio",
    "NPAMI": "Nacional | Pastilla | Mingitorio",
    "EPARE": "Exportación | Pastilla | Redonda/Flor",
    "MPARE": "Maquila | Pastilla | Redonda",
    "NPARE": "Nacional | Pastilla | Redonda",
    "NPAVP": "Nacional | Pastilla | Vestido PDCB",
    "NPAPP": "Nacional | Pastilla | ? (Camiseta PDCB)",
    "NPACR": "Nacional | Pastilla | Cristales (Moth Balls)",
    "EPCCL": "Exportación | Pastilla | Cloro (In-Tank)",
    "MPCCL": "Maquila | Pastilla | Cloro",
    "NPCCL": "Nacional | Pastilla | Cloro",
    "ECABA": "Exportación | Camphor | Barra (Round Block)",
    "ELILT": "Exportación | Líquido | Lavatrastes",
    "MLILT": "Maquila | Líquido | Lavatrastes",
    "NLILT": "Nacional | Líquido | Lavatrastes",
    "MLILD": "Maquila | Líquido | Limpieza/Desinfectante (spray)",
    "MLILS": "Maquila | Líquido | Limpiador Sanitario",
    "NLILS": "Nacional | Líquido | Limpiador Sanitario",
    "ELILS": "Exportación | Líquido | Limpiador Sanitario",
    "ETAAS": "Exportación | Tapete | Anti-Salpicadura",
    "MTAAS": "Maquila | Tapete | Anti-Salpicadura",
    "NTAAS": "Nacional | Tapete | Anti-Salpicadura",
    "ETACP": "Exportación | Tapete | Con Pastilla",
    "MTACP": "Maquila | Tapete | Con Pastilla",
    "NTACP": "Nacional | Tapete | Con Pastilla",
    "ETALI": "Exportación | Tapete | Liso",
    "MTALI": "Maquila | Tapete | Liso",
    "NTALI": "Nacional | Tapete | Liso",
    "ETAST": "Exportación | Tapete | Storm",
    "MTAST": "Maquila | Tapete | Storm",
    "NTAST": "Nacional | Tapete | Storm",
    "ENPCA": "Exportación | Sanitario | Pastilla/Cage (Rim Cage)",
    "ENPMI": "Exportación | Sanitario | Pastilla Mingitorio (Toss Block)",
    "ENPMIII": "Exportación | Sanitario | Pastilla Mingitorio III (?)",
    "ECAPR": "Exportación | Catálogo | Productos",
    "EMUFR": "Exportación | Muestrario | Fragancias",
    "MLILAV": "Maquila | Líquido | Lavatrastes (variante)",
    "NALAM": "Nacional | Alambre (para Líquido Goteador)",
    "NMANG": "Nacional | Manguera (para dispensador LG)",
    "NPIAA": "Nacional | Pilas | AA",
    "NPICC": "Nacional | Pilas | C",
    "NPLTA": "Nacional | Plástico | Tapa-Coladera",
    "NARI": "Nacional | ? | Actuador de plástico",
    "NPI": "Nacional | ? | Wick para dispensador",
    "NMPI": "Nacional | Materia Prima | ? (Front Cover)",
    "EIFIE": "? | ? | ? (registro de prueba)",
    "ELAHE": "Exportación | Lámina | Hojalata Electrolítica",
}

for pref, data in sorted(comerciales.items()):
    sig = significados.get(pref, "?")
    ej = data["ejemplos"][0]
    print(f"{pref:<14} {data['count']:>5}  {sig:<50} {ej['code']} - {ej['name'][:40]}")

print("\n" + "=" * 110)
print("MATERIALES / MATERIA PRIMA / EMPAQUE / REFACCIONES")
print("=" * 110)
print(f"{'PREFIJO':<14} {'CANT':>5}  {'SIGNIFICADO PROPUESTO':<50} {'EJEMPLO'}")
print("-" * 110)

mat_sigs = {
    "ME": "Material Empaque",
    "MP": "Materia Prima",
    "MPI": "Materia Prima Importación",
    "MS": "Material/Suministro (varios: aerosoles DC, refacciones)",
    "RP": "Refacción/Pieza (tapas, válvulas, envases)",
    "FMPI": "? | Materia Prima (GL-210A Kit)",
    "FMIP": "? | Materia Prima (GL-210A Kit) - variante",
    "FMIPI": "? | Materia Prima (GL-210A Kit) - variante",
    "MPIDA": "Materia Prima Importación | Dispensador Automático",
    "MPIVA": "Materia Prima Importación | Válvula",
    "EMPI": "Exportación | Materia Prima Importación (Keys, parts)",
    "MPEXH": "Materia Prima | Exhibidor",
}

for pref, data in sorted(materiales.items()):
    sig = mat_sigs.get(pref, "?")
    ej = data["ejemplos"][0]
    print(f"{pref:<14} {data['count']:>5}  {sig:<50} {ej['code']} - {ej['name'][:40]}")

print("\n" + "=" * 110)
print("MUESTRAS (MUE*)")
print("=" * 110)
print(f"{'PREFIJO':<14} {'CANT':>5}  {'EJEMPLO'}")
print("-" * 110)
for pref, data in sorted(muestras.items()):
    ej = data["ejemplos"][0]
    print(f"{pref:<14} {data['count']:>5}  {ej['code']} - {ej['name'][:60]}")

print("\n" + "=" * 110)
print("OTROS / HERRAMIENTAS / SERVICIOS / NO CLASIFICADOS")
print("=" * 110)
print(f"{'PREFIJO':<14} {'CANT':>5}  {'EJEMPLO'}")
print("-" * 110)
for pref, data in sorted(otros.items()):
    ej = data["ejemplos"][0]
    print(f"{pref:<14} {data['count']:>5}  {ej['code']} - {ej['name'][:60]}")

print(f"\n\nRESUMEN:")
print(f"  Comerciales: {len(comerciales)} prefijos")
print(f"  Materiales:  {len(materiales)} prefijos")
print(f"  Muestras:    {len(muestras)} prefijos")
print(f"  Otros:       {len(otros)} prefijos")
print(f"  TOTAL:       {len(prefijos)} prefijos únicos")
